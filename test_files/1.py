from openpyxl import Workbook, load_workbook

wb = load_workbook('1.FORD COURIER.xlsx')
ws = wb.active

ws['A2'] = 'xxx'
print(wb.sheetnames)


def odometer_status():
    x = input("Stan licznika na końcu miesiąca")
    ws['E8'] = int(x)
    y = input("Stan ")
    ws['E10'] = int(y)


choose = print(input('Wybierz miesiac'))
# sheets = wb.sheetnames
# ws = wb[sheets[0]] #liczba okresla ktory sheet jest aktywny // #active means last opened sheet
if choose == '1':
    ws['G1'] = 'dzila'
    odometer_status()
elif choose == 2:
    ws['G1'] = 'NIEdzila'
    odometer_status()

'''
def months():
    choose = print(input('Wybierz miesiac'))
    sheets = wb.sheetnames
    ws = wb[sheets[1]]
    if choose == 1:
        ws = wb.get_sheet_by_name[0]


i = 0
while True:
    i += 1
    if i == 3: #liczba wykonywan kodu
        break
    x = input("Stan licznika na końcu miesiąca")
    ws['E9'] = int(x)
    y = input("Stan ")
    ws['E11'] = int(y)
'''

# months()
#odometer_status()

wb.save('TEST.xlsx')

'''
9/10/18
'''
