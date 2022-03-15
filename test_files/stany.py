from openpyxl import Workbook, load_workbook

wb = load_workbook('CITROEN JUMPER_MICHAL.xlsx')
ws = wb.active

stan_styczen = input('Stan licznika na poczatku stycznia: ')
ws['E7'] = int(stan_styczen)
stan_koniec_miesiaca = input('Stan licznika na koniec miesiaca: ')
ws['E8'] = int(stan_koniec_miesiaca)
stan_paliwa_ubiegly_miesiac = input('Stan paliwa z ubieglego miesiaca: ')
ws['E10'] = int(stan_paliwa_ubiegly_miesiac)


wb.save('CITROEN JUMPER_MICHAL_1.xlsx')