import os
import glob
import re
from natsort import natsorted
import months
#import skoda
#import start
from openpyxl import Workbook, load_workbook

class DisplayMenu:
    def years_sort(self):
        path = os.getcwd()
        '''
        for fileName_relative in glob.glob(path + "**/20**", recursive=True):
            fileName_absolute = os.path.basename(fileName_relative)
            print(fileName_absolute)
        '''

        folders = ([name for name in os.listdir(path) if os.path.isdir(os.path.join(path, name))])
        folders.reverse()
        p = re.compile(r".*(venv|.idea)")
        print([p.sub("", x) for x in folders])

    def choose_year(self):
        choose = input("Choose year from 1.2022, 2.2023: ")
        if choose == "1":
            os.chdir("/Users/btn/PycharmProjects/pojazdy/2022")
            t.cars_sort()
        if choose == "2":
            os.chdir("/Users/btn/PycharmProjects/pojazdy/2023")
            t.cars_sort()

    def cars_sort(self):
        path = (os.getcwd())
        print(path)
        files = [f for f in glob.glob(path + "**/*.xlsx", recursive=True)]
        p = re.compile(r".*(/2022/|/2023/)")
        s = natsorted([p.sub('', x).strip() for x in files])
        print(s)
        choose = input("Select vehicle number: ")
        if choose == '1':
            print("You selected: ", s[0])
            wb = load_workbook('1.FORD COURIER.xlsx')
            months.ford_completing()
        if choose == '2':
            print("You selected: ", s[1])
            wb = load_workbook("2.SKODA FABIA.xlsx")
            months.skoda_completing()
        if choose == '3':
            print("You selected: ", s[2])
            wb = load_workbook("3.SKODA FABIA 2.xlsx")
            months.skoda_2_completing()
        if choose == '4':
            print("You selected: ", s[3])
            wb = load_workbook("4.FIAT PANDA.xlsx")
            months.fiat_completing()
            #t.choose_car()
        if choose == '5':
            print("You selected: ", s[4])
            wb = load_workbook("5.CITROEN JUMPER.xlsx")
            months.citroen_completing()
        if choose == '6':
            print("You selected: ", s[5])
            wb = load_workbook("6.DAEWOO LUBLIN.xlsx")
            months.daewoo_completing()
        if choose == '7':
            print("You selected: ", s[6])
            wb = load_workbook("7.FORD TRANSIT.xlsx")
            months.ford_2_completing()
        if choose == '8':
            print("You selected: ", s[7])
            wb = load_workbook("8.FORD TRANSIT LEASING.xlsx")
            months.ford_leasing_completing()
        if choose == '9':
            print("You selected: ", s[8])
            wb = load_workbook("9.FARMTRAC.xlsx")
            months.farmtrac_completing()
        if choose == '10':
            print("You selected: ", s[9])
            wb = load_workbook("10.URSUS.xlsx")
            months.ursus_completing()
        if choose == '11':
            print("You selected: ", s[10])
            wb = load_workbook("11.TYM.xlsx")
            months.tym_completing()
        if choose == '12':
            print("You selected: ", s[11])
            wb = load_workbook("12.UNIMOG.xlsx")
            months.unimog_completing()
        if choose == '13':
            print("You selected: ", s[11])
            wb = load_workbook("13.UNIMOG_2.xlsx")
            months.unimog_2_completing()
        if choose == '14':
            print("You selected: ", s[11])
            wb = load_workbook("14.NOREMAT.xlsx")
            months.noremat_completing()
        #t.choose_car()

    def choose_car(self):
        os.chdir("/Users/btn/PycharmProjects/pojazdy")
        print(os.getcwd())
        #os.system('start.py')
        months.ford_completing()


t = DisplayMenu()
#t.years_sort()                                                                      #OK
t.choose_year()
#t.cars_sort()                                                                      #OK
#t.choose_car()


'''
    REGEX
        p = re.compile(r'.*(venv|.idea)')
        print([p.sub('', x).strip() for x in folders])
        print(folders)
    
      USTAWIANIE WYLISTOWANIA TYLKO CYFR/LAT
      char_seq = "1234567890"
      for char in char_seq:
          esc_set = "*" + glob.escape(char) + "*"
          for file in glob.glob(esc_set):
              print(file)



year = input("Wybierz rok: ")
print(int(year))
if year == '2022':
    path = os.getcwd()
    cars = [name for name in os.listdir(path) if not os.path.isdir(os.path.join(path, name))]
    print(cars)
    print(os.getcwd()) #sprawdzenie w jakim katalogu sie aktualnie znajdujemy
    os.chdir(path)     #wchodzimy do katalogu 'path'
    #f = open(cars[1])
    #print(f.read())
    #f.close()
    print(os.getcwd())
    #subprocess.call(['libreoffice', 'citroen.xlsx'])

else:
    path = "/home/btn/PycharmProjects/pojazdy/2023/"
    cars = [name for name in os.listdir(path) if not os.path.isdir(os.path.join(path, name))]
    print([name for name in os.listdir(path) if not os.path.isdir(os.path.join(path, name))])
    print(cars)
    print(os.getcwd()) #sprawdzenie w jakim katalogu sie aktualnie znajdujemy
    os.chdir(path)     #wchodzimy do katalogu 'path'
    f = open(cars[0])
    print(f.read())
    f.close()



LISTOWANIE PLIKOW Z ROZSZERZENIEM xlsx
# return all files as a list
for file in os.listdir(r"/home/btn/PycharmProjects/pojazdy/"):
    # check the files which are end with specific extension
    if file.endswith(".xlsx"):
        # print path name of selected files
        print([os.path.join(file)])


REGEX
lst =  ['Paris, 458 boulevard Saint-Germain', 'Marseille, 29 rue Camille Desmoulins', 'Marseille, 1 chemin des Aubagnens']
p = re.compile(r'.*(?:rue|boulevard|quai|chemin)')
print([p.sub('', x).strip() for x in lst])

'''
