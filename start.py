from openpyxl import load_workbook
import os

path = input("Podaj sciezke do 2022: ")
os.chdir(path)

wb = load_workbook("1.FORD COURIER.xlsx")
wb1 = load_workbook("2.SKODA FABIA.xlsx")
wb2 = load_workbook("3.SKODA FABIA 2.xlsx")
wb3 = load_workbook("4.FIAT PANDA.xlsx")
wb4 = load_workbook("5.CITROEN JUMPER.xlsx")
wb5 = load_workbook("6.DAEWOO LUBLIN.xlsx")
wb6 = load_workbook("7.FORD TRANSIT.xlsx")
wb7 = load_workbook("8.FORD TRANSIT LEASING.xlsx")
wb8 = load_workbook("9.FARMTRAC.xlsx")
wb9 = load_workbook("10.URSUS.xlsx")
wb10 = load_workbook("11.TYM.xlsx")
wb11 = load_workbook("12.UNIMOG.xlsx")
wb12 = load_workbook("13.UNIMOG 2.xlsx")
wb13 = load_workbook("14.NOREMAT.xlsx")
